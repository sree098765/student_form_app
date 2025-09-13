from flask import Flask, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = "submissions.xlsx"

@app.route("/")
def index():
    with open("form.html", "r", encoding="utf-8") as f:
        return f.read()

@app.route("/submit", methods=["POST"])
def submit():
    event_name = request.form.get("event_name")
    student_name = request.form.get("student_name")
    college_email = request.form.get("college_email")

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Event Name", "Student Name", "College Email"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([event_name, student_name, college_email])
    wb.save(EXCEL_FILE)

    return "✅ Submission Saved Successfully!"

# 👉 New dashboard route
@app.route("/dashboard")
def dashboard():
    if not os.path.exists(EXCEL_FILE):
        return "No data yet!"
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    table = "<h2>Submissions</h2><table border='1' cellpadding='5'>"
    for row in rows:
        table += "<tr>" + "".join(f"<td>{cell}</td>" for cell in row) + "</tr>"
    table += "</table>"

    return table

if __name__ == "__main__":
    app.run(debug=True)
